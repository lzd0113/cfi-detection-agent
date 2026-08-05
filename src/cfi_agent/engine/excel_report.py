import os
import sys
import subprocess


def ensure_openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        return Workbook, Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("openpyxl 未安装，正在自动安装...")
        subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl'], check=True)
        print("openpyxl 安装完成")
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        return Workbook, Font, PatternFill, Alignment, Border, Side


def generate_excel(summary, modules, output_dir, include_calls=True):
    Workbook, Font, PatternFill, Alignment, Border, Side = ensure_openpyxl()
    excel_file = os.path.join(output_dir, 'cfi_function_report.xlsx')
    function_detail = summary.get('function_detail', True)

    wb = Workbook()

    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    green_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    red_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    orange_fill = PatternFill(start_color='FED7AA', end_color='FED7AA', fill_type='solid')
    blue_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    dark_fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
    owner_fill = PatternFill(start_color='FFFBEB', end_color='FFFBEB', fill_type='solid')
    title_font = Font(name='微软雅黑', size=16, bold=True, color='1E40AF')
    subtitle_font = Font(name='微软雅黑', size=11, color='64748B')

    def style_header_row(ws, row, ncols):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def style_data_cell(cell, fill=None, align=None):
        cell.border = thin_border
        if fill:
            cell.fill = fill
        cell.alignment = align or cell_align

    def fill_for_rate(rate):
        if rate == 100:
            return green_fill
        elif rate >= 70:
            return blue_fill
        elif rate >= 30:
            return orange_fill
        elif rate > 0:
            return red_fill
        else:
            return dark_fill

    ws1 = wb.active
    ws1.title = '总体统计'
    ws1.merge_cells('A1:B1')
    ws1['A1'] = 'OpenHarmony CFI 检测总体统计'
    ws1['A1'].font = title_font
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 36

    total_funcs = summary['total_cfi_protected_funcs'] + summary['total_truly_unprotected']
    protect_rate = round(summary['total_cfi_protected_funcs'] / (summary['total_cfi_protected_funcs'] + summary['total_truly_unprotected']) * 100, 1) if (summary['total_cfi_protected_funcs'] + summary['total_truly_unprotected']) > 0 else 0
    so_rate = round(summary['cfi_enabled_so'] / summary['total_so'] * 100, 1) if summary['total_so'] > 0 else 0

    stats = [
        ('指标', '数量'),
        ('.so 文件总数', summary['total_so']),
        ('CFI 已开启 .so', summary['cfi_enabled_so']),
        ('CFI 未开启 .so', summary['cfi_not_enabled_so']),
        ('.so 级 CFI 覆盖率', f'{so_rate}%'),
    ]
    if function_detail:
        stats += [
            ('CFI 保护函数', summary['total_cfi_protected_funcs']),
            ('真正未保护函数', summary['total_truly_unprotected']),
            ('函数总数', total_funcs),
            ('函数级 CFI 保护率', f'{protect_rate}%'),
        ]
    if function_detail and include_calls:
        stats += [
            ('vcall 调用点', summary.get('total_vcall_sites', 0)),
            ('vcall 有CFI', summary.get('total_vcall_cfi', 0)),
            ('vcall 无CFI', summary.get('total_vcall_no_cfi', 0)),
            ('vcall 保护率', f'{summary.get("total_vcall_cfi_rate", 0)}%'),
            ('icall 调用点', summary.get('total_icall_sites', 0)),
            ('icall 有CFI', summary.get('total_icall_cfi', 0)),
            ('icall 无CFI', summary.get('total_icall_no_cfi', 0)),
            ('icall 保护率', f'{summary.get("total_icall_cfi_rate", 0)}%'),
        ]
    if function_detail and summary.get('aarch64_so', 0) > 0:
        pac_total = summary.get('total_pac_func_protected', 0) + summary.get('total_pac_func_sign_only', 0) + summary.get('total_pac_func_no_pac', 0)
        pac_rate = round(summary.get('total_pac_func_protected', 0) / pac_total * 100, 1) if pac_total > 0 else 0
        bti_total = summary.get('total_bti_func_with', 0) + summary.get('total_bti_func_without', 0)
        bti_rate = round(summary.get('total_bti_func_with', 0) / bti_total * 100, 1) if bti_total > 0 else 0
        stats += [
            ('PAC 签名指令', summary.get('total_pac_sign_count', 0)),
            ('PAC 认证指令', summary.get('total_pac_auth_count', 0)),
            ('PAC 函数保护', summary.get('total_pac_func_protected', 0)),
            ('PAC 函数未保护', summary.get('total_pac_func_no_pac', 0)),
            ('PAC 函数覆盖率', f'{pac_rate}%'),
            ('BTI 指令数', summary.get('total_bti_count', 0)),
            ('BTI 函数覆盖', summary.get('total_bti_func_with', 0)),
            ('BTI 函数覆盖率', f'{bti_rate}%'),
        ]
    stats.append(('预编译厂商库 (无CFI)', 4))

    for row_idx, (label, value) in enumerate(stats, start=3):
        ws1.cell(row=row_idx, column=1, value=label)
        ws1.cell(row=row_idx, column=2, value=value)
        if row_idx == 3:
            style_header_row(ws1, row_idx, 2)
        else:
            style_data_cell(ws1.cell(row=row_idx, column=1), align=left_align)
            style_data_cell(ws1.cell(row=row_idx, column=2), align=cell_align)
            if '未' in label or '无' in label:
                ws1.cell(row=row_idx, column=2).fill = red_fill
            elif '已' in label or '保护' in label and '未' not in label:
                ws1.cell(row=row_idx, column=2).fill = green_fill

    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 18
    explain_row = 3 + len(stats) + 1
    ws1.merge_cells(f'A{explain_row}:B{explain_row}')
    ws1[f'A{explain_row}'] = ('LLVM CFI 机制: 每个需保护的函数生成两个版本 — func.cfi (CFI插桩入口) + func (原始函数体)。'
        'CFI 函数总数 = CFI保护函数 + CFI未保护函数，不含 .cfi 后缀函数和 __cfi* 基础设施函数，'
        '而PAC/BTI 函数总数统计包含__cfi* 基础设施函数。')
    ws1[f'A{explain_row}'].font = subtitle_font
    ws1[f'A{explain_row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws1.row_dimensions[explain_row].height = 60

    ws2 = wb.create_sheet('模块与.so详情')
    ws2.sheet_properties.outlinePr.summaryBelow = False

    if function_detail:
        base_headers = ['模块', '说明', '.so 文件路径', 'CFI状态', '.so总数', 'CFI开启', 'CFI未开启', '.so覆盖率', 'CFI保护函数', 'CFI未保护函数', '函数总数', '函数保护率', 'CFI基础设施函数']
        call_headers = ['vcall调用点', 'vcall有CFI', 'vcall无CFI', 'vcall保护率', 'icall调用点', 'icall有CFI', 'icall无CFI', 'icall保护率'] if include_calls else []
        pac_bti_headers = ['PAC签名', 'PAC认证', 'PAC覆盖率', 'BTI指令', 'BTI覆盖率']
        merged_headers = base_headers + call_headers + pac_bti_headers + ['负责人']
    else:
        merged_headers = ['模块', '说明', '.so 文件路径', 'CFI状态', '.so总数', 'CFI开启', 'CFI未开启', '.so覆盖率', '负责人']
    for col, h in enumerate(merged_headers, 1):
        ws2.cell(row=1, column=col, value=h)
    style_header_row(ws2, 1, len(merged_headers))
    ws2.row_dimensions[1].height = 36

    module_font = Font(name='微软雅黑', size=11, bold=True, color='1E40AF')
    module_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
    so_font = Font(name='微软雅黑', size=10, color='475569')

    row_idx = 2
    for m in sorted(modules, key=lambda x: -x['truly_unprotected_count'] if function_detail else -x['cfi_not_enabled_so']):
        if function_detail:
            total_funcs_mod = m['cfi_protected_count'] + m['truly_unprotected_count']
            func_rate_mod = round(m['cfi_protected_count'] / (m['cfi_protected_count'] + m['truly_unprotected_count']) * 100, 1) if (m['cfi_protected_count'] + m['truly_unprotected_count']) > 0 else 0
            module_row = [m['module'], m['desc'], '', '', m['total_so'], m['cfi_enabled_so'], m['cfi_not_enabled_so'], f'{m["cfi_rate_percent"]}%', m['cfi_protected_count'], m['truly_unprotected_count'], total_funcs_mod, f'{func_rate_mod}%', m.get('cfi_infra_count', 0)]
            if include_calls:
                module_row += [m['vcall_site_count'], m['vcall_cfi_count'], m['vcall_no_cfi_count'], f'{m["vcall_cfi_rate"]}%', m['icall_site_count'], m['icall_cfi_count'], m['icall_no_cfi_count'], f'{m["icall_cfi_rate"]}%', '']
            else:
                module_row += ['']
            pac_total_mod = m.get('pac_func_protected', 0) + m.get('pac_func_sign_only', 0) + m.get('pac_func_no_pac', 0)
            pac_rate_mod = round(m.get('pac_func_protected', 0) / pac_total_mod * 100, 1) if pac_total_mod > 0 else 0
            bti_total_mod = m.get('bti_func_with', 0) + m.get('bti_func_without', 0)
            bti_rate_mod = round(m.get('bti_func_with', 0) / bti_total_mod * 100, 1) if bti_total_mod > 0 else 0
            # Replace the owner '' at end with PAC/BTI data + new owner ''
            module_row[-1] = m.get('pac_sign_count', 0)
            module_row += [m.get('pac_auth_count', 0), f'{pac_rate_mod}%', m.get('bti_count', 0), f'{bti_rate_mod}%', '']
        else:
            module_row = [m['module'], m['desc'], '', '', m['total_so'], m['cfi_enabled_so'], m['cfi_not_enabled_so'], f'{m["cfi_rate_percent"]}%' if m['total_so'] > 0 else '0%', '']
        for col, val in enumerate(module_row, 1):
            cell = ws2.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.font = module_font
            cell.fill = module_fill
            cell.alignment = left_align if col <= 2 else cell_align
        ws2.cell(row=row_idx, column=8 if function_detail else 8).fill = fill_for_rate(m['cfi_rate_percent'])
        row_idx += 1

        so_start_row = row_idx
        for so in m['so_files']:
            path = so['path']
            if so.get('cfi_enabled'):
                cfi_status = 'CFI'
                if function_detail:
                    cfi_count = so.get('cfi_protected_count', 0)
                    unprot_count = so.get('truly_unprotected_count', 0)
                    total = cfi_count + unprot_count
                    func_rate = round(cfi_count / (cfi_count + unprot_count) * 100, 1) if (cfi_count + unprot_count) > 0 else 0
                    rate_fill = fill_for_rate(func_rate) if func_rate > 0 else dark_fill
                    so_row = [m['module'], '', path, cfi_status, '', '', '', '', cfi_count, unprot_count, total, f'{func_rate}%', so.get('cfi_infra_count', 0)]
                    if include_calls:
                        so_vcall_rate = round(so.get('vcall_cfi_count', 0) / so.get('vcall_site_count', 1) * 100, 1) if so.get('vcall_site_count', 0) > 0 else 0
                        so_icall_rate = round(so.get('icall_cfi_count', 0) / so.get('icall_site_count', 1) * 100, 1) if so.get('icall_site_count', 0) > 0 else 0
                        so_row += [so.get('vcall_site_count', 0), so.get('vcall_cfi_count', 0), so.get('vcall_no_cfi_count', 0), f'{so_vcall_rate}%', so.get('icall_site_count', 0), so.get('icall_cfi_count', 0), so.get('icall_no_cfi_count', 0), f'{so_icall_rate}%']
                    so_pac_total = so.get('pac_func_protected', 0) + so.get('pac_func_sign_only', 0) + so.get('pac_func_no_pac', 0)
                    so_pac_rate = round(so.get('pac_func_protected', 0) / so_pac_total * 100, 1) if so_pac_total > 0 else 0
                    so_bti_total = so.get('bti_func_with', 0) + so.get('bti_func_without', 0)
                    so_bti_rate = round(so.get('bti_func_with', 0) / so_bti_total * 100, 1) if so_bti_total > 0 else 0
                    so_row += [so.get('pac_sign_count', 0), so.get('pac_auth_count', 0), f'{so_pac_rate}%', so.get('bti_count', 0), f'{so_bti_rate}%', '']
                else:
                    so_row = [m['module'], '', path, cfi_status, '', '', '', '', '']
            else:
                cfi_status = 'NO CFI'
                rate_fill = dark_fill
                if function_detail:
                    cfi_count_n = so.get('cfi_protected_count', 0)
                    unprot_count_n = so.get('truly_unprotected_count', 0)
                    total_n = cfi_count_n + unprot_count_n
                    so_row = [m['module'], '', path, cfi_status, '', '', '', '', cfi_count_n, unprot_count_n, total_n, '0%', so.get('cfi_infra_count', 0)]
                    if include_calls:
                        so_row += [0, 0, 0, '0%', 0, 0, 0, '0%']
                    so_pac_total_n = so.get('pac_func_protected', 0) + so.get('pac_func_sign_only', 0) + so.get('pac_func_no_pac', 0)
                    so_pac_rate_n = round(so.get('pac_func_protected', 0) / so_pac_total_n * 100, 1) if so_pac_total_n > 0 else 0
                    so_bti_total_n = so.get('bti_func_with', 0) + so.get('bti_func_without', 0)
                    so_bti_rate_n = round(so.get('bti_func_with', 0) / so_bti_total_n * 100, 1) if so_bti_total_n > 0 else 0
                    so_row += [so.get('pac_sign_count', 0), so.get('pac_auth_count', 0), f'{so_pac_rate_n}%', so.get('bti_count', 0), f'{so_bti_rate_n}%', '']
                else:
                    so_row = [m['module'], '', path, cfi_status, '', '', '', '', '']
            for col, val in enumerate(so_row, 1):
                cell = ws2.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border
                cell.font = so_font
                cell.alignment = left_align if col in (1, 3) else cell_align
            ws2.cell(row=row_idx, column=4).fill = green_fill if so.get('cfi_enabled') else red_fill
            if function_detail:
                ws2.cell(row=row_idx, column=12).fill = rate_fill
            owner_col = (27 if include_calls else 19) if function_detail else 9
            ws2.cell(row=row_idx, column=owner_col).fill = owner_fill
            ws2.row_dimensions[row_idx].outline_level = 1
            row_idx += 1

        if so_start_row <= row_idx - 1:
            ws2.row_dimensions.group(so_start_row, row_idx - 1, outline_level=1, hidden=False)

    from openpyxl.utils import get_column_letter
    if function_detail:
        merged_widths = [20, 14, 50, 10, 10, 10, 10, 12, 14, 16, 12, 12, 12, 12, 10, 10, 12, 10, 12, 10, 12, 10, 10, 8, 10, 8, 15]
    else:
        merged_widths = [20, 14, 50, 10, 10, 10, 10, 12, 15]
    for i, w in enumerate(merged_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = 'A2'

    ws4 = wb.create_sheet('无CFI的.so')
    ws4.sheet_properties.outlinePr.summaryBelow = False

    ws4.merge_cells('A1:D1')
    ws4['A1'] = f'CFI 未开启的 .so 文件 (共 {summary["cfi_not_enabled_so"]} 个)'
    ws4['A1'].font = title_font
    ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws4.row_dimensions[1].height = 30

    ws4.cell(row=2, column=1, value='模块')
    ws4.cell(row=2, column=2, value='说明')
    ws4.cell(row=2, column=3, value='.so 文件路径')
    ws4.cell(row=2, column=4, value='负责人')
    style_header_row(ws4, 2, 4)

    row_idx = 3
    no_cfi_modules = [m for m in modules if m['cfi_not_enabled_so'] > 0]
    for m in sorted(no_cfi_modules, key=lambda x: -x['cfi_not_enabled_so']):
        module_row = [m['module'], m['desc'], f'共 {m["cfi_not_enabled_so"]} 个无CFI .so', '']
        for col, val in enumerate(module_row, 1):
            cell = ws4.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.font = module_font
            cell.fill = module_fill
            cell.alignment = left_align if col <= 2 else cell_align
        row_idx += 1

        so_start_row = row_idx
        for so in m['so_files']:
            if not so.get('cfi_enabled'):
                so_row = [m['module'], '', so['path'], '']
                for col, val in enumerate(so_row, 1):
                    cell = ws4.cell(row=row_idx, column=col, value=val)
                    cell.border = thin_border
                    cell.font = so_font
                    cell.alignment = left_align if col in (1, 3) else cell_align
                ws4.cell(row=row_idx, column=3).fill = red_fill
                ws4.cell(row=row_idx, column=4).fill = owner_fill
                ws4.row_dimensions[row_idx].outline_level = 1
                row_idx += 1

        if so_start_row <= row_idx - 1:
            ws4.row_dimensions.group(so_start_row, row_idx - 1, outline_level=1, hidden=False)

    ws4.column_dimensions['A'].width = 22
    ws4.column_dimensions['B'].width = 18
    ws4.column_dimensions['C'].width = 60
    ws4.column_dimensions['D'].width = 15
    ws4.freeze_panes = 'A3'

    ws5 = wb.create_sheet('预编译厂商库')
    ws5.merge_cells('A1:D1')
    ws5['A1'] = '预编译厂商库 (全部无 CFI)'
    ws5['A1'].font = title_font
    ws5['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws5.row_dimensions[1].height = 30

    vendor_headers = ['文件名', '大小', '功能', 'CFI状态', '负责人']
    for col, h in enumerate(vendor_headers, 1):
        ws5.cell(row=2, column=col, value=h)
    style_header_row(ws5, 2, len(vendor_headers))

    vendor_libs = [
        ('libmali-bifrost-g52-g2p0-ohos.so', '40MB', 'Mali GPU driver', 'NO CFI'),
        ('librga.z.so', '63KB', 'RGA 2D graphics', 'NO CFI'),
        ('librockchip_mpp.z.so', '1.1MB', 'Media processing', 'NO CFI'),
        ('librkaiq.z.so', '3.3MB', 'ISP image processing', 'NO CFI'),
    ]
    for row_idx, (name, size, func, status) in enumerate(vendor_libs, start=3):
        for col, val in enumerate([name, size, func, status, ''], 1):
            cell = ws5.cell(row=row_idx, column=col, value=val)
            style_data_cell(cell, align=left_align if col in (1, 3) else cell_align, fill=red_fill if col <= 4 else owner_fill)

    for i, w in enumerate([35, 10, 25, 10, 15], 1):
        ws5.column_dimensions[get_column_letter(i)].width = w

    wb.save(excel_file)
    print(f"Excel 报告已保存: {excel_file}")
    return excel_file
